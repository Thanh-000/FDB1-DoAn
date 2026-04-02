import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('tracker_fraud_v29_final.xlsx')
ws = wb['TASK_BOARD']

# The year is 2026 according to the tracker CONFIG and dates (e.g. 2026-04-05).
def parse_date(date_str):
    if not date_str or not isinstance(date_str, str):
        return date_str
    
    parts = date_str.split('/')
    if len(parts) == 2:
        day = int(parts[0])
        month = int(parts[1])
        return datetime(2026, month, day)
    return date_str

# Grab number_format from column J row 8 (a valid date cell)
valid_date_format = ws.cell(row=8, column=10).number_format

for r in range(7, 45):
    # Fix J (10) and K (11)
    for c in [10, 11]:
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, str) and '/' in cell.value:
            try:
                cell.value = parse_date(cell.value)
                cell.number_format = valid_date_format
            except Exception as e:
                print(f"Failed to parse {cell.value} at row {r} col {c}: {str(e)}")

wb.save('tracker_fraud_v29_final.xlsx')
print("✅ Fixed Date Formats to proper Excel serials.")
