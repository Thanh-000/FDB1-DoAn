import openpyxl

wb = openpyxl.load_workbook('tracker_fraud_v29_final.xlsx')
ws = wb['TASK_BOARD']

# We need to find the exactly 3 malformed rows and fix their columns.
# They are recognizable by having a datetime object or date string in Column F (Deliverable)
# Or by Task name string.
for r in range(7, 45):
    task_name = ws.cell(row=r, column=5).value
    if task_name is None:
        continue
    
    if "Lập trình cơ chế Missing-View" in str(task_name):
        ws.cell(row=r, column=6).value = "Missing-View src code"
        ws.cell(row=r, column=7).value = "TV2"
        ws.cell(row=r, column=8).value = None
        ws.cell(row=r, column=9).value = "TV3"
        ws.cell(row=r, column=10).value = "05/04"
        ws.cell(row=r, column=11).value = "12/04"
        
    elif "Thực thi 19-Experiment Ablation Matrix" in str(task_name):
        ws.cell(row=r, column=6).value = "Ablation results matrix"
        ws.cell(row=r, column=7).value = "TV4"
        ws.cell(row=r, column=8).value = "TV3"
        ws.cell(row=r, column=9).value = "TV1"
        ws.cell(row=r, column=10).value = "20/04"
        ws.cell(row=r, column=11).value = "25/04"
        
    elif "Xây dựng logic giám sát hiện tượng trôi dạt dữ liệu (Concept Drift - PSI)" in str(task_name):
        ws.cell(row=r, column=6).value = "Drift alert & narrative"
        ws.cell(row=r, column=7).value = "TV2"
        ws.cell(row=r, column=8).value = "TV4"
        ws.cell(row=r, column=9).value = "TV1"
        ws.cell(row=r, column=10).value = "27/04"
        ws.cell(row=r, column=11).value = "30/04"

wb.save('tracker_fraud_v29_final.xlsx')
print("Fixed columns for the 3 missing tasks.")
